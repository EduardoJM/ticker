from datetime import datetime
from functools import update_wrapper
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from django.urls import reverse
from django.shortcuts import redirect
from django.db import transaction, router
from django.contrib import admin
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
from django.template.response import TemplateResponse
from django.utils.translation import gettext as _
from domains.tickers.models import Ticker
from .models import PaymentEvent
from .forms import ImportPaymentEventsForm
from domains.admin.filters.autocomplete_related import (
    AutocompleteRelatedFilter,
    AutocompleteRelatedFilterMixin,
)

csrf_protect_m = method_decorator(csrf_protect)

@admin.register(PaymentEvent)
class PaymentEventAdmin(AutocompleteRelatedFilterMixin, admin.ModelAdmin):
    list_display = ['id', 'ticker', 'date', 'event_type', 'quantity', 'unit_price', 'net_value']
    list_display_links = ['id', 'ticker', 'date', 'event_type', 'quantity', 'unit_price', 'net_value']
    list_filter = [
        ("ticker", AutocompleteRelatedFilter),
    ]

    def get_urls(self):
        from django.urls import path

        def wrap(view):
            def wrapper(*args, **kwargs):
                return self.admin_site.admin_view(view)(*args, **kwargs)

            wrapper.model_admin = self
            return update_wrapper(wrapper, view)

        info = self.opts.app_label, self.opts.model_name
        
        return [
            path('import', wrap(self.import_view), name="%s_%s_import" % info)
        ] + super().get_urls()
    
    def _parse_payment_events_page(self, page: Worksheet):
        if page["A1"].value != "Produto":
            raise Exception(_('Invalid file. A1 cell should equals "Produto".'))
        if page["B1"].value != "Pagamento":
            raise Exception(_('Invalid file. B1 cell should equals "Pagamento".'))
        if page["C1"].value != "Tipo de Evento":
            raise Exception(_('Invalid file. C1 cell should equals "Tipo de Evento".'))
        if page["E1"].value != "Quantidade":
            raise Exception(_('Invalid file. E1 cell should equals "Quantidade".'))
        if page["F1"].value != "Preço unitário":
            raise Exception(_('Invalid file. F1 cell should equals "Preço unitário".'))
        if page["G1"].value != "Valor líquido":
            raise Exception(_('Invalid file. G1 cell should equals "Valor líquido".'))
        
        for row in page.rows:
            prd_cell, date_cell, type_cell, _, count_cell, price_cell, net_cell = row
            if prd_cell.value == "Produto":
                continue
            ticker = str(prd_cell.value).split('-').pop(0).strip()
            if not ticker:
                continue
            
            date = datetime.strptime(date_cell.value, "%d/%m/%Y")
            
            payment_type = None
            if type_cell.value == "Rendimento":
                payment_type = PaymentEvent.TYPE_INCOME
            if type_cell.value == "Juros Sobre Capital Próprio":
                payment_type = PaymentEvent.TYPE_INTEREST_ON_EQUITY
            if type_cell.value == "Dividendo":
                payment_type = PaymentEvent.TYPE_DIVIDEND
            
            if not payment_type:
                continue

            quantity = int(str(count_cell.value))
            unit_price = float(str(price_cell.value))
            net_value = float(str(net_cell.value))

            ticker_instance, _ = Ticker.objects.get_or_create_updated(ticker)
            PaymentEvent.objects.update_or_create(
                ticker=ticker_instance,
                event_type=payment_type,
                date=date,
                defaults={
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'net_value': net_value,
                }
            )

    def _parse_payment_events(self, file: Workbook):
        sheetnames = file.sheetnames
        for sheet_name in sheetnames:
            sheet = file[sheet_name]
            self._parse_payment_events_page(sheet)

    def _import_view(self, request, extra_context=None):
        if request.method == 'POST':
            form = ImportPaymentEventsForm(
                data=request.POST,
                files=request.FILES
            )
            if form.is_valid():
                file = form.files.get('file')
                workbook = load_workbook(file)
                self._parse_payment_events(workbook)

                info = self.opts.app_label, self.opts.model_name
                return redirect(reverse("admin:%s_%s_changelist" % info))
        else:
            form = ImportPaymentEventsForm()

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            **(extra_context or {}),
        }
        return TemplateResponse(
            request,
            "payment_events/import_payment_events.html",
            context,
        )

    @csrf_protect_m    
    def import_view(self, request, extra_context=None):
        with transaction.atomic(using=router.db_for_write(self.model)):
            return self._import_view(request, extra_context)
